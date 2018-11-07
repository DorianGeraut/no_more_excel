#!/usr/bin/python
import sys
import json
import re
from openpyxl import load_workbook, Workbook


def xlsx_to_json(ini="in.xlsx",out="out.json",start_index="A0",nb_columns=0,nb_items=0):
    """This function will read a set of lines
       as a set of object while using the
       first line as object's fields
    """
    wb = load_workbook(ini)
    ws = wb.active
    result = []
    start_x = i_to_xy(start_index)[0]
    start_y = i_to_xy(start_index)[1]

    #reading the first line:
    fields=[]
    for x in range(start_x,nb_columns+start_x):
        fields.append(ws[ xy_to_i(x, start_y)].value)

    #reading everything else:
    items = []
    for y in range(start_y+1,nb_items+start_y):
        item = {}
        for x in range(start_x,nb_columns+start_x):
            item[fields[x-start_x]] = ws[ xy_to_i(x, y)].value
        items.append(item)

    f = open(out, 'w')
    f.write(json.dumps(items,indent=3))
    f.close()

def json_to_xlsx(duplicate_from=None,ini="in.json",out="out.xlsx",start_index="A0",nb_columns=0,nb_items=0):
    """This function will write json
    object to a xlsx file as a
    set of lines using the
    first line as object's fields
    """
    f = open(ini, 'r')
    items = json.load(f)
    f.close()

    if duplicate_from:
        wb  = load_workbook(duplicate_from)
    else:
        wb = Workbook()
    ws = wb.active

    start_x = i_to_xy(start_index)[0]
    start_y = i_to_xy(start_index)[1]

    #writing the first line:
    fields = []
    x = 0
    for key in items[0].keys():
        ws[ xy_to_i(start_x+x,start_y) ] = key
        fields.append(key)
        x += 1

    #writing everything else:
    y = 1
    for item in items:
        x = 0
        for key in fields:
            print("writing "+item[key]+" in "+xy_to_i(start_x+x,start_y+y))
            ws[ xy_to_i(start_x+x,start_y+y) ] = item[key]
            x += 1
        y += 1
    wb.save(out)


def i_to_xy(index):
    num = re.compile(r"[0-9]+")
    let = re.compile(r"[A-Z]+")
    letters = let.search(index).group()
    x = 0
    for i in range(len(letters)):
        x += 27**i*(ord(letters[::-1][i])-ord('A')+1)
    y = int(num.search(index).group())
    return x,y


def xy_to_i(x, y):
    pow = 0
    while 27**(pow) < x:
        pow += 1
    tmpx = x
    letters = ""
    for p in range(pow)[::-1]:
        i = 1
        while 27**p*i <= tmpx-27**p:
            i += 1
        tmpx -= 27**p*i
        letters += chr(64+i)
    numbers = str(y)
    return letters+numbers

if __name__ == '__main__':
    print("i_to_xy(AA11) = "+i_to_xy("AA11"))
