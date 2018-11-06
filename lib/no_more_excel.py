#!/usr/bin/python
import sys
import json
from openpyxl import load_workbook, Workbook


def xlsx_to_json(ini="in.xlsx",out="out.json",start_index="A0",nb_columns=0,nb_items=0):
    """This function will read a set of lines
       as a set of object while using the
       first line as object's fields
    """
    wb = load_workbook(ini)
    ws = wb.active
    result = []
    start_x = i_to_xy(start_index)[0]
    start_y = i_to_xy(start_index)[1]

    #reading the first line:
    fields=[]
    for x in range(start_x,nb_columns+start_x):
        fields.append(ws[ xy_to_i(x, start_y)].value)

    #reading everything else:
    items = []
    for y in range(start_y+1,nb_items+start_y):
        item = {}
        for x in range(start_x,nb_columns+start_x):
            item[fields[x-start_x]] = ws[ xy_to_i(x, y)].value
        items.append(item)

    f = open(out, 'w')
    f.write(json.dumps(items,indent=3))
    f.close()

def json_to_xlsx(duplicate,ini="in.json",out="out.xlsx",start_index="A0",nb_columns=0,nb_items=0):
    """This function will write json
    object to a xlsx file as a
    set of lines using the
    first line as object's fields
    """
    f = open(ini, 'r')
    items = json.load(f)
    f.close()

    if duplicate:
        wb  = load_workbook(duplicate)
    else:
        wb = Workbook()
    ws = wb.active

    start_x = i_to_xy(start_index)[0]
    start_y = i_to_xy(start_index)[1]

    #writing the first line:
    fields = []
    x = 0
    for key in items[0].keys():
        ws[ xy_to_i(start_x+x,start_y) ] = key
        fields.append(key)
        x += 1

    #writing everything else:
    y = 1
    for item in items:
        x = 0
        for key in fields:
            print("writing "+item[key]+" in "+xy_to_i(start_x+x,start_y+y))
            ws[ xy_to_i(start_x+x,start_y+y) ] = item[key]
            x += 1
        y += 1
    wb.save(out)


def i_to_xy(index):
    x = ord(index[0])-ord('A')
    y = int(index[1])
    return x,y


def xy_to_i(x, y):
    index = ""
    index += chr(ord('A')+x)
    index += str(y)
    return index

if __name__ == '__main__':
    xlsx_to_json("exemple2.xlsx","exemple2.json",start_index="B6",nb_columns=5,nb_items=7)
    json_to_xlsx("exemple3.json","exemple3.xlsx",start_index="B6",nb_columns=5,nb_items=7)
